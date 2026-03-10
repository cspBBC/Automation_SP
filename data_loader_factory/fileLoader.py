"""Test Data Loaders - Load test data from various formats (JSON, CSV, Excel)."""

import os
import json
import csv
import logging
from pathlib import Path
from typing import Dict, Any, List
from abc import ABC, abstractmethod

logger = logging.getLogger('sp_validation')


class BaseLoader(ABC):
    """Abstract base class for data loaders."""
    
    @staticmethod
    @abstractmethod
    def load(file_path: str) -> Dict[str, Any]:
        """Load data from file and return as dictionary.
        
        Args:
            file_path: Path to the data file
            
        Returns:
            Dictionary containing loaded data
        """
        pass


class JSONLoader(BaseLoader):
    """Load test data from JSON files."""
    
    @staticmethod
    def load(file_path: str) -> Dict[str, Any]:
        """Load test data from a JSON file.
        
        Args:
            file_path: Path to the JSON file
            
        Returns:
            Dictionary containing test data keyed by SP name
        """
        # Ensure .json extension
        if not file_path.endswith('.json'):
            file_path = f"{file_path}.json"
        
        # Build path if not absolute
        if not os.path.isabs(file_path):
            project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            
            # Try multiple paths in order
            candidate_paths = [
                os.path.join(project_root, file_path),
                os.path.join(project_root, 'data_layer', 'test_data', file_path),
            ]
            
            file_path = None
            for path in candidate_paths:
                if os.path.exists(path):
                    file_path = path
                    break
            
            if not file_path:
                # None found, use first candidate for error message
                file_path = candidate_paths[0]
        
        logger.info(f"Loading JSON test data from: {file_path}")
        
        if not os.path.isfile(file_path):
            raise FileNotFoundError(f"Test data file not found: {file_path}")
        
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            logger.info(f"Successfully loaded test data from {file_path}")
            return data
        except json.JSONDecodeError as e:
            logger.error(f"JSON parsing error in {file_path}: {e}")
            raise
        except Exception as e:
            logger.error(f"Error loading JSON file: {e}")
            raise


class CSVLoader(BaseLoader):
    """Load test data from CSV files with automatic schema detection.
    
    Auto-detects schema and supports both:
    - Keyword-driven format (Module/Operation/Test Case ID/etc.)
    - Generic format (sp_name)
    """
    
    @staticmethod
    def load(file_path: str) -> Dict[str, Any]:
        """Load test data from a CSV file with auto-detected schema.
        
        Detects whether CSV uses keyword-driven format (Module/Operation/Test Case ID/etc.)
        or generic format (sp_name) and parses accordingly.
        
        Args:
            file_path: Path to the CSV file
            
        Returns:
            Dictionary containing test data keyed by module/SP name
        """
        # Ensure .csv extension
        if not file_path.endswith('.csv'):
            file_path = f"{file_path}.csv"
        
        # Build path if not absolute
        if not os.path.isabs(file_path):
            project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            
            # Try multiple paths in order
            candidate_paths = [
                os.path.join(project_root, file_path),
                os.path.join(project_root, 'data_layer', 'test_data', file_path),
            ]
            
            file_path = None
            for path in candidate_paths:
                if os.path.exists(path):
                    file_path = path
                    break
            
            if not file_path:
                # None found, use first candidate for error message
                file_path = candidate_paths[0]
        
        logger.info(f"Loading CSV test data from: {file_path}")
        
        if not os.path.isfile(file_path):
            raise FileNotFoundError(f"Test data file not found: {file_path}")
        
        try:
            data = {}
            
            with open(file_path, 'r', encoding='utf-8-sig') as f:
                # Read header to detect schema
                reader = csv.DictReader(f)
                headers = reader.fieldnames or []
                
                # Detect schema: keyword-driven or generic
                is_keyword_driven = 'Module' in headers and 'Operation' in headers
                
                logger.info(f"Detected schema: {'keyword-driven' if is_keyword_driven else 'generic'}")
                
                for row in reader:
                    if is_keyword_driven:
                        # Keyword-driven schema parsing
                        module = row.get('Module', '').strip()
                        operation = row.get('Operation', '').strip()
                        case_id = row.get('Test Case ID', '').strip()
                        executed = row.get('Executed', 'No').strip().lower() == 'yes'
                        test_type = row.get('Test Type', 'independent').strip()
                        params_json = row.get('test_parameters', '{}').strip()
                        
                        # Extract validation fields from CSV (expected_status, expected_result, etc.)
                        expected_status = row.get('expected_status', '').strip()
                        expected_result = row.get('expected_result', '').strip()
                        test_description = row.get('test_description', '').strip()
                        
                        # Skip if module is empty
                        if not module:
                            logger.warning(f"Skipping row with empty Module: {row}")
                            continue
                        
                        # Initialize module entry if not exists
                        if module not in data:
                            data[module] = []
                        
                        # Parse test parameters JSON
                        try:
                            params = json.loads(params_json) if params_json else {}
                        except json.JSONDecodeError as e:
                            logger.warning(f"Invalid JSON in test_parameters for {case_id}: {e}")
                            params = {}
                        
                        # Build test case object matching keyword-driven format
                        test_case = {
                            'case_id': case_id,
                            'case_type': operation.upper() if operation else 'POSITIVE',
                            'description': f"{operation} test case: {case_id}",
                            'operation': operation,
                            'executed': executed,
                            'test_type': test_type,
                            'parameters': params,
                            'expected_status': expected_status,
                            'expected_result': expected_result,
                            'expected_message_pattern': expected_result,  # Map expected_result to expected_message_pattern for validation
                            'test_description': test_description
                        }
                        
                        data[module].append(test_case)
                    else:
                        # Generic schema parsing (sp_name)
                        sp_name = row.get('sp_name', 'unknown')
                        if sp_name not in data:
                            data[sp_name] = []
                        data[sp_name].append(row)
            
            logger.info(f"Successfully loaded test data from {file_path}")
            return data
        except csv.Error as e:
            logger.error(f"CSV parsing error in {file_path}: {e}")
            raise
        except Exception as e:
            logger.error(f"Error loading CSV file: {e}")
            raise


class ExcelLoader(BaseLoader):
    """Load test data from Excel files."""
    
    @staticmethod
    def load(file_path: str) -> Dict[str, Any]:
        """Load test data from an Excel file.
        
        Supports both generic format (sp_name) and keyword-driven format
        (Module/Operation/Test Case ID/Executed/Test Type/test_parameters).
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Dictionary containing test data keyed by module name
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required to load Excel files. Install it with: pip install openpyxl")
        
        # Ensure .xlsx/.xls extension
        if not file_path.endswith(('.xlsx', '.xls')):
            file_path = f"{file_path}.xlsx"
        
        # Build path if not absolute
        if not os.path.isabs(file_path):
            project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            
            # Try multiple paths in order
            candidate_paths = [
                os.path.join(project_root, file_path),
                os.path.join(project_root, 'data_layer', 'test_data', file_path),
            ]
            
            file_path = None
            for path in candidate_paths:
                if os.path.exists(path):
                    file_path = path
                    break
            
            if not file_path:
                # None found, use first candidate for error message
                file_path = candidate_paths[0]
        
        logger.info(f"Loading Excel test data from: {file_path}")
        
        if not os.path.isfile(file_path):
            raise FileNotFoundError(f"Test data file not found: {file_path}")
        
        try:
            data = {}
            wb = openpyxl.load_workbook(file_path)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                headers = [cell.value for cell in ws[1]]
                
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                    row_dict = dict(zip(headers, row))
                    
                    # Support both generic format (sp_name) and keyword-driven format (Module/Operation/etc)
                    if 'Module' in row_dict:
                        # Keyword-driven format
                        module = row_dict.get('Module', '').strip() if row_dict.get('Module') else ''
                        operation = row_dict.get('Operation', '').strip() if row_dict.get('Operation') else ''
                        case_id = row_dict.get('Test Case ID', '').strip() if row_dict.get('Test Case ID') else ''
                        executed = str(row_dict.get('Executed', 'No')).strip().lower() == 'yes'
                        test_type = row_dict.get('Test Type', 'independent').strip() if row_dict.get('Test Type') else 'independent'
                        params_json = row_dict.get('test_parameters', '{}').strip() if row_dict.get('test_parameters') else '{}'
                        
                        if not module:
                            logger.warning(f"Skipping row with empty Module at row {row_idx}")
                            continue
                        
                        if module not in data:
                            data[module] = []
                        
                        # Parse test parameters JSON
                        try:
                            params = json.loads(params_json) if params_json else {}
                        except (json.JSONDecodeError, TypeError) as e:
                            logger.warning(f"Invalid JSON in test_parameters for {case_id}: {e}")
                            params = {}
                        
                        # Build test case object matching keyword-driven format
                        test_case = {
                            'case_id': case_id,
                            'case_type': operation.upper() if operation else 'POSITIVE',
                            'description': f"{operation} test case: {case_id}",
                            'operation': operation,
                            'executed': executed,
                            'test_type': test_type,
                            'parameters': params
                        }
                        data[module].append(test_case)
                    else:
                        # Generic format using sp_name
                        sp_name = row_dict.get('sp_name', 'unknown')
                        if sp_name not in data:
                            data[sp_name] = []
                        data[sp_name].append(row_dict)
            
            logger.info(f"Successfully loaded test data from {file_path}")
            return data
        except Exception as e:
            logger.error(f"Error loading Excel file: {e}")
            raise

