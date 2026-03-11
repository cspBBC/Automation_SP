"""Excel Loader - Load test data from Excel files."""

import os
import logging
import json
from typing import Dict, Any
from .base_loader import BaseLoader

logger = logging.getLogger('sp_validation')


class ExcelLoader(BaseLoader):
    """Load test data from Excel files.
    
    Requires openpyxl or xlrd package to be installed.
    """
    
    @staticmethod
    def load(file_path: str) -> Dict[str, Any]:
        """Load test data from an Excel file.
        
        Args:
            file_path: Path to the Excel file (with or without .xlsx/.xls extension)
            
        Returns:
            Dictionary containing test data keyed by SP name
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required to load Excel files. Install it with: pip install openpyxl")
        
        # Ensure file path has .xlsx extension if not specified
        if not file_path.endswith(('.xlsx', '.xls')):
            file_path = f"{file_path}.xlsx"
        
        # Check if file exists in data_layer/test_data/ directory
        if not os.path.isabs(file_path):
            project_root_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
            file_path = os.path.join(project_root_dir, 'data_layer', 'test_data', file_path)
        
        logger.info(f"Loading Excel test data from: {file_path}")
        
        if not os.path.isfile(file_path):
            raise FileNotFoundError(f"Test data file not found: {file_path}")
        
        try:
            data = {}
            wb = openpyxl.load_workbook(file_path)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                headers = [cell.value for cell in ws[1]]
                
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                    row_dict = dict(zip(headers, row))
                    
                    # Support both generic format (sp_name) and keyword-driven format (Module/Operation/Test Case ID)
                    if 'Module' in row_dict:
                        # Keyword-driven format (same as CSV loader)
                        module = row_dict.get('Module', '').strip() if row_dict.get('Module') else ''
                        operation = row_dict.get('Operation', '').strip() if row_dict.get('Operation') else ''
                        case_id = row_dict.get('Test Case ID', '').strip() if row_dict.get('Test Case ID') else ''
                        executed = str(row_dict.get('Executed', 'No')).strip().lower() == 'yes'
                        test_type = row_dict.get('Test Type', 'independent').strip() if row_dict.get('Test Type') else 'independent'
                        params_json = row_dict.get('test_parameters', '{}').strip() if row_dict.get('test_parameters') else '{}'
                        
                        if not module:
                            logger.warning(f"Skipping row with empty Module at row {row_idx}")
                            continue
                        
                        if module not in data:
                            data[module] = []
                        
                        # Parse test parameters JSON
                        try:
                            params = json.loads(params_json) if params_json else {}
                        except (json.JSONDecodeError, TypeError) as e:
                            logger.warning(f"Invalid JSON in test_parameters for {case_id}: {e}")
                            params = {}
                        
                        # Build test case object matching keyword-driven CSV format
                        test_case = {
                            'case_id': case_id,
                            'case_type': operation.upper() if operation else 'POSITIVE',
                            'description': f"{operation} test case: {case_id}",
                            'operation': operation,
                            'executed': executed,
                            'test_type': test_type,
                            'parameters': params
                        }
                        data[module].append(test_case)
                    else:
                        # Generic format using sp_name
                        sp_name = row_dict.get('sp_name', 'unknown')
                        if sp_name not in data:
                            data[sp_name] = []
                        data[sp_name].append(row_dict)
            
            logger.info(f"Successfully loaded test data from {file_path}")
            return data
        except Exception as e:
            logger.error(f"Error loading Excel file: {e}")
            raise
